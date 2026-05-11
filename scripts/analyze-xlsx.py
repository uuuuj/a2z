"""사용자템플릿_엑셀_SHI.xlsx 구조 분석 (REQ-002/T-012 검증용)"""
import openpyxl
import sys

import sys
p = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\duddl\Desktop\Project\a2z\사용자템플릿_엑셀.xlsx'
wb = openpyxl.load_workbook(p, data_only=False)

print('=' * 80)
print('SHEET LIST')
print('=' * 80)
for s in wb.sheetnames:
    ws = wb[s]
    print(f'  - {s}  ({ws.max_row} rows x {ws.max_column} cols)')

for s in wb.sheetnames:
    ws = wb[s]
    print()
    print('=' * 80)
    print(f'SHEET: {s}')
    print('=' * 80)
    print(f'Dimensions: {ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}')
    print(f'Page: orientation={ws.page_setup.orientation} paperSize={ws.page_setup.paperSize}')
    print(f'Print area: {ws.print_area}')
    print()
    print(f'-- Merged cells ({len(ws.merged_cells.ranges)}) --')
    for r in sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
        c = ws.cell(row=r.min_row, column=r.min_col)
        val = c.value
        print(f'  {r}  =>  {repr(val) if val is not None else "EMPTY"}')

    print()
    print('-- Non-empty cells --')
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and str(c.value).strip():
                print(f'  {c.coordinate}: {repr(c.value)}')

    print()
    n_img = len(ws._images) if hasattr(ws, '_images') else 0
    n_chart = len(ws._charts) if hasattr(ws, '_charts') else 0
    print(f'-- Images: {n_img}, Charts: {n_chart} --')
    if n_img > 0:
        for i, img in enumerate(ws._images):
            anchor = img.anchor
            try:
                print(f'  image[{i}] anchor={anchor._from.col},{anchor._from.row}')
            except Exception:
                print(f'  image[{i}] anchor={anchor}')

    print()
    print('-- Column widths (custom, px) --')
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width is not None:
            print(f'  {col_letter}: width={dim.width}')

    print('-- Row heights (custom) --')
    for row_num, dim in ws.row_dimensions.items():
        if dim.height is not None:
            print(f'  row {row_num}: height={dim.height}')
