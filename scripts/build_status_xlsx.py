# -*- coding: utf-8 -*-
"""GitHub 이슈를 원천으로 개발 현황 Excel과 JSON 백업을 생성한다.

실행
  python scripts/build_status_xlsx.py               # Excel + JSON 둘 다 생성
  python scripts/build_status_xlsx.py --json-only   # 백업만
  python scripts/build_status_xlsx.py --out 경로.xlsx

이슈가 정본이고 Excel은 **읽기 전용 생성물**이다. Excel을 손으로 고치지 말 것 —
다음 실행에서 덮어써진다. 상태를 바꾸려면 이슈의 `상태:` 라벨을 고친다.

JSON 백업은 이슈가 GitHub 서버에만 있는 문제를 덮는다. 저장소에 파일로 남겨
clone만으로도 이력을 읽을 수 있게 한다.
"""
import io, os, re, json, subprocess, argparse, datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, '개발 현황.xlsx')
DUMP = os.path.join(ROOT, 'docs/tracking-archive/issues.json')

STATUS_ORDER = ['상태: 개발 중', '상태: 실기 확인', '상태: API 대기', '상태: API 요청 필요',
                '상태: 분석 필요', '상태: 개발 대기']
DRAW_ORDER = ['도면: 제작도', '도면: 조립도', '도면: 설치도', '도면: 가공도', '도면: 공통']

# 사내 이슈 관리 화면으로 내보내는 항목 — 서비스하려면 반드시 해결해야 하는 것.
# 주체 라벨은 둘 다 붙을 수 있다 (예: 우리가 정의하면서 소프트힐스도 같이 보는 건).
MUST = '필수'
MUST_ROLES = [('필수: API 개발', 'API 개발'), ('필수: 내부 해결', '내부 해결')]
CLOSED_STATES = ('완료', '미채택')


def fetch():
    """이슈 전체 — 본문·코멘트·라벨·타임스탬프까지."""
    fields = 'number,title,state,stateReason,labels,body,comments,createdAt,closedAt,updatedAt,url'
    r = subprocess.run(['gh', 'issue', 'list', '--state', 'all', '--limit', '1000',
                        '--json', fields], capture_output=True, text=True, encoding='utf-8')
    if r.returncode != 0:
        raise SystemExit('gh 조회 실패: %s' % (r.stderr or '')[:300])
    return json.loads(r.stdout or '[]')


def labels(i):
    return [x['name'] for x in i['labels']]


def pick(i, order, prefix):
    ls = labels(i)
    for want in order:
        if want in ls:
            return want.split(': ', 1)[1]
    for l in ls:
        if l.startswith(prefix):
            return l.split(': ', 1)[1]
    return ''


def meta(i):
    """이관·백필 때 심어둔 메타 블록에서 원본 정보를 읽는다."""
    out = {}
    for pat in (r'<!-- excel-meta(.*?)-->', r'<!-- tracking-meta(.*?)-->'):
        m = re.search(pat, i.get('body') or '', re.S)
        if not m:
            continue
        for line in m.group(1).split('\n'):
            if ':' in line:
                k, v = line.split(':', 1)
                out.setdefault(k.strip(), v.strip())
    return out


def clean_title(t):
    return re.sub(r'^(?:\[[^\]]+\]\s*)+', '', t).strip()


def rows(iss):
    out = []
    for i in sorted(iss, key=lambda x: x['number']):
        m = meta(i)
        closed = i['state'] == 'CLOSED'
        ls = labels(i)
        out.append({
            '번호': i['number'],
            '제목': clean_title(i['title']),
            '상태': ('완료' if i.get('stateReason') != 'NOT_PLANNED' else '미채택') if closed
                    else (pick(i, STATUS_ORDER, '상태: ') or '미분류'),
            '도면': pick(i, DRAW_ORDER, '도면: '),
            '생산 필수': 'O' if '생산 필수' in ls else '',
            '필수': 'O' if MUST in ls else '',
            '해결 주체': '+'.join(s for full, s in MUST_ROLES if full in ls),
            '대분류': m.get('대분류', ''),
            '원본 ID': m.get('id', '') or (('Excel No.' + m['no']) if m.get('no') else ''),
            '최초 구현일': m.get('최초구현', ''),
            '완료 확인일': m.get('완료확인', '') or (i.get('closedAt') or '')[:10],
            '생성': (i.get('createdAt') or '')[:10],
            '최근 갱신': (i.get('updatedAt') or '')[:10],
            '코멘트': len(i.get('comments') or []),
            '라벨': ', '.join(l for l in ls
                            if l != MUST and not l.startswith(('상태: ', '도면: ', '필수: '))),
            'URL': i.get('url', ''),
        })
    return out


def must_open(rs):
    """사내 화면으로 내보낼 항목 — 열려 있는 `필수`만. 개발사 몫을 먼저 둔다."""
    svc = [r for r in rs if r['필수'] == 'O' and r['상태'] not in CLOSED_STATES]
    return sorted(svc, key=lambda r: (0 if 'API 개발' in r['해결 주체'] else 1, r['번호']))


def summary(rs):
    tot = len(rs)
    done = sum(1 for r in rs if r['상태'] == '완료')
    skip = sum(1 for r in rs if r['상태'] == '미채택')
    live = tot - skip
    must = [r for r in rs if r['생산 필수'] == 'O']
    must_done = sum(1 for r in must if r['상태'] == '완료')
    by = {}
    for r in rs:
        if r['상태'] not in ('완료', '미채택'):
            by[r['상태']] = by.get(r['상태'], 0) + 1
    svc = must_open(rs)
    lines = [('전체 이슈', tot), ('완료', done), ('미채택', skip),
             ('유효 관리 대상', live), ('진행 중', live - done),
             ('진행률', ('%.1f%%' % (done * 100.0 / live)) if live else '-'),
             ('', ''),
             ('생산 필수 전체', len(must)), ('생산 필수 완료', must_done),
             ('생산 필수 잔여', len(must) - must_done),
             ('', ''),
             ('필수 잔여', len(svc)),
             ('필수 · 개발사', sum(1 for r in svc if 'API 개발' in r['해결 주체'])),
             ('필수 · 내부', sum(1 for r in svc if '내부 해결' in r['해결 주체'])),
             ('', '')]
    for k in sorted(by, key=lambda x: -by[x]):
        lines.append(('진행 중 · ' + k, by[k]))
    return lines


def write_xlsx(rs, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()

    # ── 현황 요약 ──
    ws = wb.active
    ws.title = '현황 요약'
    ws['A1'] = '개발 현황 (GitHub 이슈 자동 생성)'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = '생성 시각 %s · 원천은 GitHub 이슈이며 이 파일은 읽기 전용 생성물입니다. 손으로 고치지 마십시오.' \
               % datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    ws['A2'].font = Font(size=9, color='7F7F7F')
    r = 4
    for k, v in summary(rs):
        if k:
            ws.cell(r, 1, k).font = Font(bold=True)
            ws.cell(r, 2, v)
        r += 1
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14

    # ── 필수 (사내 보고) ──
    #   사내 이슈 관리 화면이 그대로 받아 쓰는 시트. 개발사 몫과 내부 몫이 한눈에 갈리게 둔다.
    #   해결 주체가 둘 다인 항목은 한 줄에 `API 개발+내부 해결`로 표기한다 (양쪽이 동시에 진행 중인 경우).
    wsm = wb.create_sheet('필수')
    mcols = ['번호', '제목', '해결 주체', '상태', '도면', '최근 갱신', 'URL']
    for c, name in enumerate(mcols, 1):
        cell = wsm.cell(1, c, name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='24292E')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    api_fill = PatternFill('solid', fgColor='FBE9E9')
    in_fill = PatternFill('solid', fgColor='E7F0F7')
    thin_m = Side(style='thin', color='D9D9D9')
    for ri, row in enumerate(must_open(rs), 2):
        for c, name in enumerate(mcols, 1):
            cell = wsm.cell(ri, c, row[name])
            cell.border = Border(bottom=thin_m)
            cell.alignment = Alignment(vertical='top', wrap_text=(name == '제목'))
        wsm.cell(ri, 3).fill = api_fill if 'API 개발' in row['해결 주체'] else in_fill
    for c, w in enumerate([7, 58, 18, 12, 10, 11, 42], 1):
        wsm.column_dimensions[get_column_letter(c)].width = w
    wsm.freeze_panes = 'A2'

    # ── 이슈 목록 ──
    ws2 = wb.create_sheet('이슈 목록')
    cols = list(rs[0].keys()) if rs else []
    head = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='2563EB')
    thin = Side(style='thin', color='D9D9D9')
    for c, name in enumerate(cols, 1):
        cell = ws2.cell(1, c, name)
        cell.font = head
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    done_fill = PatternFill('solid', fgColor='EAFAF4')
    must_fill = PatternFill('solid', fgColor='FDECEC')
    for ri, row in enumerate(rs, 2):
        for c, name in enumerate(cols, 1):
            cell = ws2.cell(ri, c, row[name])
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical='top', wrap_text=(name == '제목'))
        if row['상태'] == '완료':
            ws2.cell(ri, 3).fill = done_fill
        elif row['생산 필수'] == 'O':
            ws2.cell(ri, 5).fill = must_fill
    width = {'번호': 7, '제목': 58, '상태': 12, '도면': 10, '생산 필수': 9, '필수': 7,
             '해결 주체': 18, '대분류': 12,
             '원본 ID': 14, '최초 구현일': 12, '완료 확인일': 12, '생성': 11,
             '최근 갱신': 11, '코멘트': 8, '라벨': 26, 'URL': 42}
    for c, name in enumerate(cols, 1):
        ws2.column_dimensions[get_column_letter(c)].width = width.get(name, 14)
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(cols)), len(rs) + 1)

    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default=XLSX)
    ap.add_argument('--json-only', action='store_true')
    a = ap.parse_args()

    iss = fetch()
    rs = rows(iss)
    print('이슈 %d건 조회' % len(iss))

    os.makedirs(os.path.dirname(DUMP), exist_ok=True)
    io.open(DUMP, 'w', encoding='utf-8', newline='\n').write(
        json.dumps(iss, ensure_ascii=False, indent=1, sort_keys=True))
    print('백업 → %s' % os.path.relpath(DUMP, ROOT).replace('\\', '/'))

    if a.json_only:
        return
    write_xlsx(rs, a.out)
    print('Excel → %s' % os.path.relpath(a.out, ROOT).replace('\\', '/'))
    for k, v in summary(rs)[:6]:
        if k:
            print('   %-14s %s' % (k, v))


if __name__ == '__main__':
    main()
