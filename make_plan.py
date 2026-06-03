# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Malgun Gothic"
HEADERS = ["단계", "주차", "기간", "세부 활동", "산출물", "상태", "마일스톤", "비고 · 리스크"]
WIDTHS  = [16, 7, 13, 46, 20, 9, 22, 28]

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def is_ex(text):
    return isinstance(text, str) and text.strip().startswith("Ex.")

def build_sheet(ws, title, goal, period, rows, theme):
    title_fill = PatternFill("solid", fgColor=theme["title"])
    head_fill  = PatternFill("solid", fgColor=theme["head"])
    stage_fill = PatternFill("solid", fgColor=theme["stage"])
    ms_fill    = PatternFill("solid", fgColor="FFF2CC")
    prog_fill  = PatternFill("solid", fgColor="FCE4D6")
    ncol = len(HEADERS); last_col = get_column_letter(ncol)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]; c.value = title
    c.font = Font(name=FONT, size=15, bold=True, color="FFFFFF"); c.fill = title_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]; c.value = "목표  ·  " + goal
    c.font = Font(name=FONT, size=10, color="404040")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    ws.merge_cells(f"A3:{last_col}3")
    c = ws["A3"]; c.value = "기간  ·  " + period
    c.font = Font(name=FONT, size=10, color="404040")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 22

    hr = 4
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=hr, column=j, value=h)
        c.font = Font(name=FONT, size=10.5, bold=True, color="FFFFFF"); c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[hr].height = 26

    start = hr + 1
    for i, r in enumerate(rows):
        rr = start + i; ws.row_dimensions[rr].height = 32
        for j, v in enumerate(r, start=1):
            c = ws.cell(row=rr, column=j, value=v); c.border = border
            align = Alignment(vertical="center", wrap_text=True,
                              horizontal="left" if j in (1, 4, 5, 7, 8) else "center")
            c.font = Font(name=FONT, size=10, italic=True, color="9A9A9A") if is_ex(v) \
                     else Font(name=FONT, size=10, color="404040")
            c.alignment = align
            if j == 1:
                c.fill = stage_fill; c.font = Font(name=FONT, size=10, bold=True, color="333333")
            if j == 7 and isinstance(v, str) and v.strip():
                c.fill = ms_fill; c.font = Font(name=FONT, size=9.5, bold=True, color="7F6000")
            if j == 6 and v == "진행":
                c.fill = prog_fill; c.font = Font(name=FONT, size=10, bold=True, color="C55A11")

    merge_start = start
    for i in range(1, len(rows) + 1):
        prev = rows[i-1][0]
        if i == len(rows) or rows[i][0] != prev:
            rr_end = start + i - 1
            if rr_end > merge_start:
                ws.merge_cells(f"A{merge_start}:A{rr_end}")
                ws[f"A{merge_start}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            merge_start = start + i

    for j, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{start}"
    ws.sheet_view.showGridLines = False


rows1 = [
 ("가공도 출력 본진","W1","6/4~6/10","가공도 페이지 배치 — 여러 부재 페이지 분할 + PDF 출력","가공도 페이지·PDF","진행","","현재 진행 중인 본진"),
 ("가공도 출력 본진","W2","6/11~6/17","자동·수동 출력 경로 통합 + 가공도 정확도(회전·EA·홀 종류)","통합 출력·정확도","예정","",""),
 ("치수·BOM·레이아웃","W3","6/18~6/24","치수 품질(겹침·보조선·외곽) + BOM 표기 + 뷰 레이아웃","2D 품질 개선","예정","◆ 1차 본진 완성","Ex. 빠른 개발로 3주 내 압축"),
 ("제작도 기능 완성","W4","6/25~7/1","각도(角度) 치수 표시 기능 추가","각도 표시","예정","",""),
 ("제작도 기능 완성","W5","7/2~7/8","No Paint(도장 제외) 영역 표시 기능 추가","No Paint 표시","예정","",""),
 ("제작도 기능 완성","W6","7/9~7/15","부재별 점선/실선 정합 + 조달 제출용 도면 정리","제출용 도면 준비","예정","◆ 제작도 핵심 기능 완성",""),
 ("사외 조달 검증","W7","7/16~7/22","조달팀 1차 제출 — 도면 보고 제작 가능 여부 의뢰","도면 제출 #1","예정","◆ 7월 중 1차 제출","7월 내 발송 목표"),
 ("사외 조달 검증","W8","7/23~7/29","Ex. 회신 대기 — 진행 중 발굴되는 추가 기능 보완(병행)","Ex. 추가 기능","예정","","Ex. 활동 더 생길 수 있음"),
 ("사외 조달 검증","W9","7/30~8/5","제작 가능 여부 회신 수집 + 문제점 분석","검토 의견 분석","예정","","Ex. 회신 일정 외부 의존"),
 ("사외 조달 검증","W10","8/6~8/12","'제작 불가·수정' 지적 사항 도면 반영","도면 보완","예정","",""),
 ("사외 조달 검증","W11","8/13~8/19","보완 도면 2차 제출 + 재검토","도면 제출 #2","예정","◆ 2차 도면 제출","Ex. 제출→보완 반복"),
 ("사외 조달 검증","W12","8/20~8/26","2차 회신 반영 + Ex. 추가 보완","2차 반영","예정","",""),
 ("사외 조달 검증","W13","8/27~9/2","최종 보완 + 지속 개선","보완·개선","예정","◆ 분기 점검",""),
]

rows2 = [
 ("분석 · 환경","W1","6/4~6/10","기존 디지털 트윈 구조·자산 파악","현황 분석","예정","",""),
 ("분석 · 환경","W2","6/11~6/17","위치추적기 데이터 규격·API 명세 확인 + Unity 개발환경 구축","요구·환경 정리","예정","◆ 착수 점검",""),
 ("실시간 데이터 연동","W3","6/18~6/24","위치추적기 API 실시간 수신 연결","API 수신 모듈","예정","",""),
 ("실시간 데이터 연동","W4","6/25~7/1","수신 데이터 파싱·구조화","파싱 모듈","예정","",""),
 ("실시간 데이터 연동","W5","7/2~7/8","좌표계 정합 (센서 좌표 ↔ 트윈 좌표)","좌표 변환","예정","","Ex. 정합 난이도에 따라 변동"),
 ("트윈 반영","W6","7/9~7/15","크레인 실시간 위치를 트윈에 반영","위치 동기화","예정","",""),
 ("트윈 반영","W7","7/16~7/22","자세(회전·붐 각도 등) 반영","자세 동기화","예정","","Ex. 수신 데이터 항목 따라 변동"),
 ("트윈 반영","W8","7/23~7/29","끊김·지연 보정, 실시간 안정화","실시간 안정화","예정","◆ 실시간 연동 1차 완성",""),
 ("기능 추가","W9","7/30~8/5","Ex. 실시간 모니터링 화면","Ex. 모니터링 기능","예정","","Ex. 기능 미확정"),
 ("기능 추가","W10","8/6~8/12","Ex. 크레인 운동 이력 재생","Ex. 이력 재생","예정","","Ex. 기능 미확정"),
 ("기능 추가","W11","8/13~8/19","Ex. 경고·알람 / 간섭·안전 체크","Ex. 경고 기능","예정","","Ex. 기능 미확정"),
 ("통합 · 검증","W12","8/20~8/26","실데이터 통합 검증","통합 검증","예정","",""),
 ("통합 · 검증","W13","8/27~9/2","데모 + 보고","데모·보고","예정","◆ 중간 보고",""),
]

rows3 = [
 ("기술 조사","W1","6/4~6/10","Ex. AM 환경·라이선스 확인","Ex.","예정","","Ex. 방향 미정"),
 ("기술 조사","W2","6/11~6/17","Ex. API/PML 등 연동 방식 조사","Ex.","예정","","Ex. 방향 미정"),
 ("(미정)","W3","6/18~6/24","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W4","6/25~7/1","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W5","7/2~7/8","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W6","7/9~7/15","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W7","7/16~7/22","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W8","7/23~7/29","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W9","7/30~8/5","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W10","8/6~8/12","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W11","8/13~8/19","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W12","8/20~8/26","Ex. 방향 확정 후 채움","","","",""),
 ("(미정)","W13","8/27~9/2","Ex. 방향 확정 후 채움","","","",""),
]

theme1 = {"title":"1F4E79", "head":"2E75B6", "stage":"DDEBF7"}
theme2 = {"title":"375623", "head":"548235", "stage":"E2EFDA"}
theme3 = {"title":"595959", "head":"808080", "stage":"EDEDED"}

wb = Workbook()
ws1 = wb.active; ws1.title = "2D 자동제작도 (VIZCore)"
ws2 = wb.create_sheet("F-C 디지털트윈 (Unity)")
ws3 = wb.create_sheet("Aveva Marine (AM)")

build_sheet(ws1, "과제 ①  VIZCore 기반 2D 자동 제작도",
    "제작도·가공도 자동 생성 고도화 → 7월 중 사외 조달팀 제작 가능 여부 검증",
    "2026-06-04(목) ~ 09-02(화)  ·  13주", rows1, theme1)
build_sheet(ws2, "과제 ②  Unity F/C(해상크레인) 디지털 트윈",
    "기존 디지털 트윈에 위치추적기 실시간 데이터를 API로 연동하고 기능 추가",
    "2026-06-04(목) ~ 09-02(화)  ·  13주", rows2, theme2)
build_sheet(ws3, "과제 ③  Aveva Marine(AM) 기반 3D 프로그램",
    "개발 방향 검토 중 — 확정 후 세부 일정 채움 (현재 틀만)",
    "2026-06-04(목) ~ 09-02(화)  ·  13주", rows3, theme3)

out = r"C:\Users\duddl\Desktop\3개월_주간_개발계획_2026.xlsx"
wb.save(out)
print("saved:", out)
