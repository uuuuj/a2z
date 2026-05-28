# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"C:\Users\duddl\Desktop\Project\a2z\가공도_로직_API_순서_체크리스트.xlsx"
KFONT = "맑은 고딕"

# (no, phase_key, item, api, purpose, vars, desc, star)
rows = [
    (1,"A","템플릿 확인","File.Exists","엑셀 템플릿 존재 확인","사용자템플릿_엑셀_가공도.xlsx","없으면 TemplateMissing → PDF 0개. 경로 맞는지",False),
    (2,"A","UI 잠금","lvDrawingSheet.Enabled","출력 중 시트 클릭 차단","false","출력 동안 목록 회색·클릭 안 됨",False),
    (3,"A","화면갱신 차단","vizcore3d.BeginUpdate()","중간 깜빡임 차단","—","50번 EndUpdate와 짝. 누락 시 격리 깜빡임",False),
    (4,"A","3D 어노테이션 초기화","Review.Note.Clear / Review.Measure.Clear / ShapeDrawing.Clear","이전 풍선·치수·보조선 제거","—","잔재 0건이어야",False),
    (5,"A","2D 초기화","Clear2DView()  (내부)","2D 캔버스 비우기","—","이전 페이지 잔상 없어야",False),
    (6,"A","X-Ray 끄기","View.XRay.Enable","투시 끄기","false","격리 시 반투명 방지",False),
    (7,"A","전체 표시","Object3D.Show","BOM 수집용 전체 표시","Object3DKind.ALL, true","—",False),
    (8,"A","선택 해제","Object3D.Select","빨간 하이라이트 제거","Object3dSelectionModes.DESELECT_ALL","—",False),
    (9,"A","렌더모드","View.SetRenderMode","은선 모드","RenderModes.DASH_LINE","—",False),
    (10,"A","BOM 표 채우기","CollectBOMInfo  (내부)","우측 BOM 8컬럼 채움","false, syntheticSheet(-3, allMfgBomIndices)","제작도와 동일 방식. 전체 부재 1회",False),
    (11,"A","BOM snapshot","SnapshotBomRows()  (내부)","live ListView 복사(race 차단)","—","행 수 = Min(부재수,15) 일치 확인",False),
    (12,"A","페이지 분할","SplitMfgIntoPages  (내부)","5부재/페이지 분할","mfgSheets, rowsPerPage=5","페이지 수 = ⌈부재수/5⌉",False),

    (13,"B","캔버스 초기화","ResetCanvasForMfgPage() → Drawing2D.View.SetCanvasSize / SetSelectCanvas","A4 가로 캔버스","297, 210  /  1","페이지마다 깨끗한 A4",False),
    (14,"B","슬롯 데이터","BuildMfgPageData  (내부)","Input_N 사전 구성","page, totalPages, struName, bomSnapshot","좌측 이름 Input_5~9, BOM Input_10~129, 미치환 {Input_N} 노출 0",False),
    (15,"B","템플릿 주입","Drawing2D.Template.ImportExcelWithData","텍스트를 템플릿에 주입","xlsxPath, data","도면정보·BOM·이름 칸 채워졌는지",False),
    (16,"B","View 영역 캐시","Drawing2D.Template.GetViewAreasFromExcel","View_1~5 좌표 1회 캐시","xlsxPath","5개 영역 모두 존재. 누락 시 throw",False),

    (17,"C","행 진입 cleanup","Review.Note.Clear / Measure.Clear / ShapeDrawing.Clear","부재 간 누적 차단","—","행마다 0에서 시작",False),
    (18,"C","부재 격리","Object3D.Show  ×2","대상 1개만 표시","(ALL,false) → ([idx],true)","화면에 1개 부재만",False),
    (19,"C","최장축 판별","(계산)","가로 배치축 결정","sizeX/Y/Z 비교 → LongestAxis","X/Y/Z 중 가장 긴 축",False),
    (20,"C","PAD/PLATE 판별","Object3D.UDA.FromIndex  (GetSprefValue)","평판 여부 → 카메라 분기",'index, "SPREF"',"PAD/PLATE면 최단축 정면",False),
    (21,"C","카메라 방향","View.MoveCamera","정면 뷰 세팅","CameraDirection.X_PLUS / Y_PLUS / Z_PLUS","부재가 정면으로",False),
    (22,"C","ORIENTATION 회전","View.RotateCameraByScreenAxis  (ApplyOrientationRotation)","UDA 각도 회전","0, 0, angle","ORIENTATION 있으면 그만큼 회전",False),
    (23,"C","Z90 결정","(계산)","세로 부재 가로화 플래그",'pose.ApplyZ90 = (LongestAxis=="Z")',"Z 최장축이면 true",False),
    (24,"C","Osnap 수집","Object3D.GetOsnapPoint","외곽 치수 기준점 수집","bom.Index","LINE 끝점·POINT만(CIRCLE 제외). 부재 양 외곽이 다 잡히는지 = 버그 핵심",True),
    (25,"C","은선 필터","FilterHiddenLineOsnap  (내부)","뒷면 점 제거","osnap, viewDir, BBox, isMinusCamera","깊이축 뒤쪽 15% 제거",False),
    (26,"C","좌표 병합","MergeCoordinates  (내부)","근접점 통합","osnap, tolerance=0.5","중복 점 합쳐짐",False),
    (27,"C","체인치수 생성","AddChainDimensionByAxis  (내부)","축별 연속 치수","merged, axis, 0.5, viewDir","입력 점이 절반이면 치수도 절반(Garbage-In/Out)",False),
    (28,"C","치수 스타일","Review.Measure.GetStyle / SetStyle","Blue·소수0·프레임 off","MeasureStyle","색·폰트 통일",False),
    (29,"C","치수 그리기","DrawDimension  (내부→Measure)","치수선 1개씩 추가","start,end,axis,offset(50/100/250),min×3,viewDir,extLines,max×3,posOff","오프셋 방향·겹침 발생 지점. 6mm/59mm 위치 확인",True),
    (30,"C","보조선 등록","ShapeDrawing.AddLine","치수 보조선","extLines, -1, Color.Blue, 0.3f, true","pose.ShapeDrawingIds에 누적",False),
    (31,"C","풍선 추가","Review.Note.GetStyle / Note.AddNoteSurface","홀·슬롯·R 풍선 4분면","text, textPos, arrowPos, NoteStyle  ×N","풍선이 모델 밖 사분면에",False),

    (32,"D","렌더모드·실루엣","View.SetRenderMode / View.SilhouetteEdge / SilhouetteEdgeColor","은선 + 윤곽 강조","DASH_LINE / true / Color.Green","—",False),
    (33,"D","카메라 fit","View.FlyToObject3d","부재 화면 채움","[bom.Index], 1.25f","부재가 적절히 큼",False),
    (34,"D","Z90/R180 회전","View.RotateCameraByScreenAxis","가로화·EA펼침","0,0,90 또는 0,0,180 (조건부)","ApplyZ90/R180일 때만",False),
    (35,"D","2D 캡처","Drawing2D.Object2D.Create2DViewObjectWithModelHiddenLineAtCanvasOrigin","현재뷰 → 2D 객체","Drawing2D_ModelViewKind.CURRENT  (선두께 2.0)","objId ≥ 0이어야 성공",False),
    (36,"D","크기 측정","Drawing2D.Object2D.GetObjectSize / GetObjectScale","fit 비율 계산","objId, ref w, ref h","NaN/0 가드",False),
    (37,"D","스케일 맞춤","Drawing2D.Object2D.RescaleObject","View 영역에 맞춤","objId, newScale = curScale × fitRatio","영역 대비 적정 크기",False),
    (38,"D","영역 중앙 배치","Drawing2D.Object2D.MoveObjectTo","View_N 중앙으로","objId, cx=area.X+W/2, cy=area.Y+H/2","슬롯 정중앙",False),
    (39,"D","보조선 2D 변환","Drawing2D.Object2D.Add2DObjectFromShapeDrawing","보조선 → 2D","pose.ShapeDrawingIds  (선 0.1, SOLID)","보조선이 PDF에",False),
    (40,"D","풍선 2D 변환","Drawing2D.View.Add2DNoteFrom3DNote","풍선 → 2D","noteIds[]  (텍스트 3.5)","풍선이 PDF에",False),
    (41,"D","치수 텍스트 시프트","ApplyParallelTextShift  (내부)","텍스트 겹침 회피","viewDir, GetObjectScale(objId), Measure.Items","소형 뷰면 skip, 중대형이면 방향오판 가능 = 겹침 표면 증상",True),
    (42,"D","치수 2D 변환","Drawing2D.Measure.Add2DMeasureFrom3DMeasure","치수 → 2D","measureIds[]  (선 0.5)","치수가 PDF에",False),
    (43,"D","행 실패 cleanup","Drawing2D.Object2D.DeleteObjectBy2DView","실패 객체 제거","objId  (finally, 실패 시만)","다음 행 오염 방지",False),

    (44,"E","2D 렌더","Drawing2D.Render","최종 렌더","—","모든 객체 반영",False),
    (45,"E","선택 해제","Drawing2D.Object2D.UnselectAllObjectBy2DView / UnselectCurrentWorkObjectBy2DView","선택 테두리 제거","—","PDF에 선택표시 없어야",False),
    (46,"E","PDF 저장","Drawing2D.Object2D.Export2PDFBy2DView","페이지 PDF 출력","pdfPath  (MakeUniquePdfPath: 이름_가공도_p1of2_S0_타임스탬프)","파일 생성·충돌 없음",False),

    (47,"F","BOM UI 복원","CollectBOMInfo  (내부)","이전 선택 시트 BOM 되돌림","false, previousSelectedSheet","출력 전 화면 상태로",False),
    (48,"F","UI 잠금 해제","lvDrawingSheet.Enabled","목록 재활성","prevLvEnabled","다시 클릭 가능",False),
    (49,"F","가시성 복원","Object3D.Show  ×2","선택 시트 부재만 격리 복원","(ALL,false) → (previousSelectedSheet.MemberIndices, true)","출력 전 미리보기 격리 유지",False),
    (50,"F","화면갱신 재개","vizcore3d.EndUpdate()","최종 상태 1회 반영","—","3번과 짝",False),
]

PHASE = {
    "A": ("A. 진입·초기화", "출력당 1회", "D9E1F2"),
    "B": ("B. 페이지 준비", "페이지당 1회", "E2EFDA"),
    "C": ("C. 3D 장면 생성", "부재당 1회", "FFF2CC"),
    "D": ("D. 2D 캡처·변환", "부재당 1회", "FCE4D6"),
    "E": ("E. 페이지 렌더·PDF", "페이지당 1회", "DDEBF7"),
    "F": ("F. 종료 복원", "출력당 1회", "EDEDED"),
}

headers = ["순서","단계","반복","항목","API","목적","API 사용 변수","설명 / 체크포인트","핵심","실제 결과","판정","비고"]
widths  = [6, 14, 11, 18, 44, 22, 40, 50, 6, 18, 9, 18]

wb = Workbook()
ws = wb.active
ws.title = "가공도_API_순서"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title (row1) + legend (row2)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
t = ws.cell(1,1,"가공도 출력 로직 API 실행 순서 체크리스트  —  수동 PDF 경로 (btnMfgDrawingSheet_Click → GenerateMfgDrawingManual)")
t.font = Font(name=KFONT, size=13, bold=True, color="FFFFFF")
t.fill = PatternFill("solid", fgColor="1F4E78")
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
lg = ws.cell(2,1,"반복 단위: A·F=출력당 1회, B·E=페이지당, C·D=부재당   |   C(3D 장면 생성)는 미리보기 경로(ExecuteMfgDrawing)와 공유   |   ★=치수 겹침 버그(P3-3) 결과 체크 핵심 (24·29·41)   |   브랜치 refactor/dead-code")
lg.font = Font(name=KFONT, size=9, italic=True, color="444444")
lg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 18

HROW = 4
for c,(h,w) in enumerate(zip(headers,widths), start=1):
    cell = ws.cell(HROW,c,h)
    cell.font = Font(name=KFONT, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[cell.column_letter].width = w
ws.row_dimensions[HROW].height = 24

star_fill = PatternFill("solid", fgColor="C00000")
for i, r in enumerate(rows):
    no, pk, item, api, purpose, vars_, desc, star = r
    plabel, prepeat, pcolor = PHASE[pk]
    excel_r = HROW + 1 + i
    fill = PatternFill("solid", fgColor=pcolor)
    vals = [no, plabel, prepeat, item, api, purpose, vars_, desc, "★" if star else "", "", "", ""]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(excel_r, c, v)
        cell.border = border
        cell.alignment = Alignment(
            horizontal=("center" if c in (1,9,11) else "left"),
            vertical="center", wrap_text=True)
        cell.font = Font(name=KFONT, size=10,
                         bold=(c==4 and star),
                         color=("C00000" if (c==4 and star) else "000000"))
        cell.fill = fill
    # 핵심 셀 강조
    if star:
        kc = ws.cell(excel_r, 9)
        kc.fill = star_fill
        kc.font = Font(name=KFONT, size=11, bold=True, color="FFFFFF")
    # 체크 입력칸(실제결과/판정/비고)은 흰 배경
    for c in (10,11,12):
        ws.cell(excel_r, c).fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[excel_r].height = 42

LAST = HROW + len(rows)
# 판정 드롭다운
dv = DataValidation(type="list", formula1='"OK,NG,보류,N/A"', allow_blank=True)
dv.add(f"K{HROW+1}:K{LAST}")
ws.add_data_validation(dv)

ws.auto_filter.ref = f"A{HROW}:L{LAST}"
ws.freeze_panes = "E5"
ws.sheet_view.showGridLines = False

wb.save(OUT)
print("SAVED:", OUT, "rows:", len(rows))
